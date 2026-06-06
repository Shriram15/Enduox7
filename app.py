from flask import Flask, jsonify, request, render_template, send_file
from werkzeug.exceptions import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import os
import logging

app = Flask(__name__)

# ================ RACE STATE ================
# States: 'setup', 'running', 'waiting', 'finished'
race = {
    "state": "setup",
    "current_hour": 1,
    "hour_start_time": None,
    "config": {
        "max_time_per_hour": 3600
    },
    "teams": {},
    "events": []
}

def get_max_laps(category):
    if category == "solo_f":
        return 12
    return 14  # duo and solo_m


def format_time(sec):
    """Format seconds into HH:MM:SS string."""
    sec = max(0, int(sec))
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h:02}:{m:02}:{s:02}"


def add_event(event_type, message, bib=None):
    """Add an event to the race event log."""
    if race["state"] == "running" and race["hour_start_time"]:
        elapsed = time.time() - race["hour_start_time"]
    else:
        elapsed = 0
    
    race["events"].append({
        "time": time.strftime("%H:%M:%S"),
        "elapsed": format_time(elapsed),
        "hour": race["current_hour"],
        "type": event_type,
        "message": message,
        "bib": bib
    })



def generate_excel(filename="race_results.xlsx"):
    """Helper to generate excel files for manual and auto exports."""
    wb = Workbook()
    
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    dnf_fill = PatternFill(start_color="FC8181", end_color="FC8181", fill_type="solid")
    ok_fill = PatternFill(start_color="68D391", end_color="68D391", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    first_sheet = True
    for bib, team in race["teams"].items():
        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()
        ws.title = f"Bib_{bib}"

        max_laps = get_max_laps(team["category"])

        c = ws.cell(row=1, column=1, value="Lap")
        c.font = hdr_font; c.fill = hdr_fill; c.border = thin_border

        hours = sorted(team["laps"].keys()) if team["laps"] else [1]

        for col, hour in enumerate(hours, start=2):
            c = ws.cell(row=1, column=col, value=f"Hour {hour}")
            c.font = hdr_font; c.fill = hdr_fill; c.border = thin_border

        for lap_num in range(1, max_laps + 1):
            ws.cell(row=lap_num + 1, column=1, value=f"Lap {lap_num}").border = thin_border

            for col, hour in enumerate(hours, start=2):
                laps = team["laps"].get(hour, [])

                if team["dnf_info"] and isinstance(team["dnf_info"], dict):
                    if team["dnf_info"].get("reason") == "timeout":
                        dnf_hour = team["dnf_info"].get("hour")
                        dnf_laps = team["dnf_info"].get("laps_completed", 0)
                        if hour == dnf_hour and lap_num == dnf_laps + 1:
                            c = ws.cell(row=lap_num + 1, column=col, value="DNF")
                            c.fill = dnf_fill; c.border = thin_border
                            continue

                if lap_num <= len(laps):
                    lap = laps[lap_num - 1]
                    text = f"{lap['runner']} | {lap['clock_time']} | {format_time(lap['lap_time'])}"
                else:
                    text = ""

                ws.cell(row=lap_num + 1, column=col, value=text).border = thin_border

        row_s = max_laps + 3
        ws.cell(row=row_s, column=1, value="Category").font = Font(bold=True)
        ws.cell(row=row_s, column=2, value=team["category"].upper())
        ws.cell(row=row_s + 1, column=1, value="Runner M Total").font = Font(bold=True)
        ws.cell(row=row_s + 1, column=2, value=format_time(team["runner_time"]["M"]))
        ws.cell(row=row_s + 2, column=1, value="Runner F Total").font = Font(bold=True)
        ws.cell(row=row_s + 2, column=2, value=format_time(team["runner_time"]["F"]))
        total = team["runner_time"]["M"] + team["runner_time"]["F"]
        ws.cell(row=row_s + 3, column=1, value="Bib Total").font = Font(bold=True)
        ws.cell(row=row_s + 3, column=2, value=format_time(total))
        ws.cell(row=row_s + 4, column=1, value="Status").font = Font(bold=True)
        sc = ws.cell(row=row_s + 4, column=2, value="DNF" if team["dnf"] else "Active")
        sc.fill = dnf_fill if team["dnf"] else ok_fill

        ws.column_dimensions["A"].width = 18
        for ci in range(2, len(hours) + 2):
            letter = chr(64 + ci) if ci <= 26 else "A"
            ws.column_dimensions[letter].width = 30

    ws_lb = wb.create_sheet("Leaderboard")
    hdrs = ["Rank", "Bib", "Category", "Total Laps", "Total Time", "Runner M", "Runner F", "Status"]
    for ci, h in enumerate(hdrs, 1):
        c = ws_lb.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = thin_border

    for rank, entry in enumerate(build_leaderboard(), 1):
        ws_lb.cell(row=rank + 1, column=1, value=rank).border = thin_border
        ws_lb.cell(row=rank + 1, column=2, value=f"Bib {entry['bib']}").border = thin_border
        ws_lb.cell(row=rank + 1, column=3, value=entry["category"]).border = thin_border
        ws_lb.cell(row=rank + 1, column=4, value=entry["laps"]).border = thin_border
        ws_lb.cell(row=rank + 1, column=5, value=entry["time_fmt"]).border = thin_border
        ws_lb.cell(row=rank + 1, column=6, value=entry["runner_m"]).border = thin_border
        ws_lb.cell(row=rank + 1, column=7, value=entry["runner_f"]).border = thin_border
        sc = ws_lb.cell(row=rank + 1, column=8, value=entry["status"])
        sc.fill = dnf_fill if entry["dnf"] else ok_fill
        sc.border = thin_border

    for letter in "ABCDEFGH":
        ws_lb.column_dimensions[letter].width = 18

    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(filepath)
    return filepath


def check_hour_state():
    """Check if the current hour has ended."""
    if race["state"] != "running" or race["hour_start_time"] is None:
        return

    now = time.time()
    elapsed = now - race["hour_start_time"]
    max_time = race["config"]["max_time_per_hour"]

    # Has the hour ended?
    if elapsed >= max_time:
        hour = race["current_hour"]

        # Check all teams for DNF
        for bib, team in race["teams"].items():
            # Skip teams that already won their category (locked) or are DNF
            if team.get("locked"):
                continue
            if team["dnf"]:
                continue
            max_laps = get_max_laps(team["category"])
            laps_completed = len(team["laps"].get(hour, []))

            # Duo-specific rule: each runner (M and F) must complete at least 2 laps in the hour
            if team.get("category") == "duo":
                laps_list = team["laps"].get(hour, [])
                m_laps = sum(1 for lap in laps_list if lap.get("runner") == "M")
                f_laps = sum(1 for lap in laps_list if lap.get("runner") == "F")
                if m_laps < 2 or f_laps < 2:
                    team["dnf"] = True
                    team["dnf_info"] = {
                        "reason": "min_runner_laps",
                        "hour": hour,
                        "m_laps": m_laps,
                        "f_laps": f_laps
                    }
                    add_event(
                        "elimination",
                        f"Bib {bib} eliminated (insufficient runner laps: M:{m_laps}, F:{f_laps} in Hour {hour})",
                        bib
                    )
                    continue

            # Timeout/full-lap check (legacy rule)
            if laps_completed < max_laps:
                team["dnf"] = True
                team["dnf_info"] = {
                    "reason": "timeout",
                    "hour": hour,
                    "laps_completed": laps_completed
                }
                add_event(
                    "elimination",
                    f"Bib {bib} eliminated ({laps_completed}/{max_laps} laps in Hour {hour})",
                    bib
                )

        # Transition state
        race["state"] = "waiting"
        
        # AUTO EXPORT
        if race["teams"]:
            generate_excel(f"race_results_Hour_{hour}.xlsx")
            add_event("system", f"Hour {hour} ended. Auto-exported results to race_results_Hour_{hour}.xlsx")

        race["current_hour"] += 1
        
        # Check if race is over (0 or 1 teams left)
        active = [bib for bib, t in race["teams"].items() if not t["dnf"]]
        if len(active) <= 1:
            race["state"] = "finished"


def get_winner_info():
    """Check if there's a winner."""
    if race["state"] == "setup":
        return None

    active = [bib for bib, t in race["teams"].items() if not t["dnf"]]

    if len(active) == 1:
        return {"winner": active[0], "finished": True}
    elif len(active) == 0:
        return {"winner": None, "finished": True, "all_eliminated": True}
    return None


def build_leaderboard():
    """Build sorted leaderboard data."""
    entries = []
    for bib, team in race["teams"].items():
        total_laps = sum(len(laps) for laps in team["laps"].values())
        total_time = team["runner_time"]["M"] + team["runner_time"]["F"]

        cat_display = "Duo"
        if team["category"] == "solo_m": cat_display = "Solo M"
        if team["category"] == "solo_f": cat_display = "Solo F"

        entries.append({
            "bib": bib,
            "category": cat_display,
            "laps": total_laps,
            "time": total_time,
            "time_fmt": format_time(total_time),
            "dnf": team["dnf"],
            "runner_m": format_time(team["runner_time"]["M"]),
            "runner_f": format_time(team["runner_time"]["F"]),
            "status": "DNF" if team["dnf"] else "Active"
        })

    entries.sort(key=lambda x: (x["dnf"], -x["laps"], x["time"]))
    return entries


def build_leaderboards():
    """Build separate leaderboards per category and mark category winners."""
    result = {"duo": [], "solo_m": [], "solo_f": []}

    for bib, team in race["teams"].items():
        total_laps = sum(len(laps) for laps in team["laps"].values())
        total_time = team["runner_time"]["M"] + team["runner_time"]["F"]

        cat = team["category"]
        cat_display = "Duo"
        if cat == "solo_m": cat_display = "Solo M"
        if cat == "solo_f": cat_display = "Solo F"

        entry = {
            "bib": bib,
            "category": cat_display,
            "laps": total_laps,
            "time": total_time,
            "time_fmt": format_time(total_time),
            "dnf": team["dnf"],
            "runner_m": format_time(team["runner_time"]["M"]),
            "runner_f": format_time(team["runner_time"]["F"]),
            "status": "DNF" if team["dnf"] else "Active",
            "winner": False
        }

        if cat in result:
            result[cat].append(entry)

    # sort each
    for k in result:
        result[k].sort(key=lambda x: (x["dnf"], -x["laps"], x["time"]))

    # determine category winners (only if sole active remains and constraints met)
    winners = get_category_winners()
    for cat_key, wbib in winners.items():
        if wbib is None:
            continue
        # mark the entry in the leaderboard and lock the team to prevent further changes
        for e in result.get(cat_key, []):
            if int(e["bib"]) == int(wbib):
                e["winner"] = True
                break
        # set locked flag on the team so it cannot be modified/eliminated later
        try:
            t = race["teams"].get(int(wbib))
            if t is not None:
                t["locked"] = True
                t["category_winner"] = True
        except Exception:
            pass

    return result


def get_category_winners():
    """Return a dict of winners per category or None if not determined."""
    winners = {"duo": None, "solo_m": None, "solo_f": None}

    for cat in winners.keys():
        active = [bib for bib, t in race["teams"].items() if t["category"] == cat and not t["dnf"]]
        if len(active) == 1:
            cand_bib = active[0]
            cand = race["teams"][cand_bib]
            # For duo require both runners to have at least 2 laps each
            if cat == "duo":
                m_laps = sum(1 for laps in cand["laps"].values() for lap in laps if lap["runner"] == "M")
                f_laps = sum(1 for laps in cand["laps"].values() for lap in laps if lap["runner"] == "F")
                if m_laps >= 2 and f_laps >= 2:
                    winners[cat] = cand_bib
                else:
                    # If one runner failed to reach minimum, mark the bib DNF
                    cand["dnf"] = True
                    cand["dnf_info"] = {"reason": "min_laps", "M_laps": m_laps, "F_laps": f_laps}
                    add_event("elimination", f"Bib {cand_bib} eliminated — duo runner requirement not met (M:{m_laps},F:{f_laps})", cand_bib)
            else:
                winners[cat] = cand_bib

    return winners


def build_status_response():
    """Build full race status response."""
    check_hour_state()
    
    elapsed = 0
    if race["state"] == "running" and race["hour_start_time"]:
        elapsed = time.time() - race["hour_start_time"]

    teams_data = {}
    for bib, team in race["teams"].items():
        disp_hour = race["current_hour"]
        if race["state"] == "finished" and disp_hour > 1:
            disp_hour -= 1
            
        laps_this_hour = len(team["laps"].get(disp_hour, []))
        
        teams_data[str(bib)] = {
            "category": team["category"],
            "target_laps": get_max_laps(team["category"]),
            "laps": {str(k): v for k, v in team["laps"].items()},
            "runners": team["runners"],
            "dnf": team["dnf"],
            "dnf_info": team["dnf_info"],
            "total_time": format_time(team["total_time"]),
            "runner_time": {
                "M": format_time(team["runner_time"]["M"]),
                "F": format_time(team["runner_time"]["F"])
            },
            "laps_this_hour": laps_this_hour,
            "total_laps": sum(len(l) for l in team["laps"].values())
        }

    winner = get_winner_info()
    if winner and race["state"] != "finished":
        race["state"] = "finished"

    return {
        "state": race["state"],
        "elapsed": format_time(elapsed),
        "elapsed_seconds": elapsed,
        "current_hour": race["current_hour"],
        "config": race["config"],
        "teams": teams_data,
        "leaderboard": build_leaderboard(),
        "leaderboards": build_leaderboards(),
        "winners_by_category": get_category_winners(),
        "events": race["events"][-100:],
        "winner": winner
    }


def category_has_winner(category):
    """Return True if the given category already has a declared winner (and ensure locks refreshed)."""
    # Only trust explicit flags set on teams; do not auto-evaluate winners here.
    for t in race["teams"].values():
        if t.get("category") == category and t.get("category_winner"):
            return True
    return False


def is_category_winner(bib):
    """Return True if the bib has already been declared a category winner."""
    winners = get_category_winners()
    for cat, wbib in winners.items():
        try:
            if wbib is not None and int(wbib) == int(bib):
                return True
        except Exception:
            continue
    return False


# ================ ROUTES ================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config", methods=["GET"])
def get_config():
    return jsonify(race["config"])


@app.route("/api/config", methods=["POST"])
def set_config():
    if race["state"] != "setup":
        return jsonify({"error": "Cannot change configuration after the race has started."}), 409

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "No data provided."}), 400

    if "max_time_per_hour" in data:
        try:
            mt = int(data["max_time_per_hour"])
            if mt < 10 or mt > 7200:
                return jsonify({"error": "Hour duration must be between 10 and 7200 seconds."}), 400
            race["config"]["max_time_per_hour"] = mt
        except (ValueError, TypeError):
            return jsonify({"error": "Hour duration must be a valid integer."}), 400

    return jsonify({"success": True, "config": race["config"]})


@app.route("/api/teams", methods=["POST"])
def add_team():
    if race["state"] != "setup":
        return jsonify({"error": "Cannot add bibs after setup is complete."}), 409

    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"error": "No data provided."}), 400

    try:
        bib = int(data.get("bib", ""))
    except (ValueError, TypeError):
        return jsonify({"error": "Bib must be a valid integer."}), 400
        
    if bib in race["teams"]:
        return jsonify({"error": f"Bib {bib} already exists."}), 409
        
    category = data.get("category", "")
    if category not in ["duo", "solo_m", "solo_f"]:
        return jsonify({"error": "Invalid category."}), 400
        
    runners_active = {"M": True, "F": True}
    if category == "solo_m": runners_active["F"] = False
    if category == "solo_f": runners_active["M"] = False
        
    race["teams"][bib] = {
        "category": category,
        "laps": {},
        "last_time_this_hour": None,
        "total_time": 0.0,
        "runners": runners_active,
        "dnf": False,
        "dnf_info": None,
        "runner_time": {"M": 0.0, "F": 0.0},
        "milestones": {"M": False, "F": False}
    }
    
    return jsonify({"success": True, "teams": race["teams"]})


@app.errorhandler(Exception)
def handle_all_exceptions(e):
    # Return HTTPExceptions with their original codes and descriptions
    if isinstance(e, HTTPException):
        return jsonify({"error": e.description}), e.code

    # Log unexpected errors and return a generic message
    logging.exception("Unhandled exception: %s", e)
    return jsonify({"error": "Internal server error."}), 500


@app.route("/api/teams/<int:bib>", methods=["DELETE"])
def remove_team(bib):
    if race["state"] != "setup":
        return jsonify({"error": "Cannot remove bibs after setup is complete."}), 409
        
    if bib in race["teams"]:
        del race["teams"][bib]
        
    return jsonify({"success": True, "teams": race["teams"]})


@app.route("/api/race/start_hour", methods=["POST"])
def start_hour():
    check_hour_state()
    
    if race["state"] == "running":
        return jsonify({"error": "An hour is already running."}), 409
    
    if race["state"] == "finished":
        return jsonify({"error": "The race is already finished."}), 409

    if race["state"] == "setup":
        if len(race["teams"]) < 2:
            return jsonify({"error": "At least 2 bibs are required to start."}), 400
        race["events"] = []

    race["state"] = "running"
    race["hour_start_time"] = time.time()
    
    for team in race["teams"].values():
        team["last_time_this_hour"] = None

    add_event("start", f"🏁 Hour {race['current_hour']} started!")
    return jsonify({"success": True, "hour": race["current_hour"]})


@app.route("/api/race/status", methods=["GET"])
def race_status():
    return jsonify(build_status_response())


@app.route("/api/lap", methods=["POST"])
def record_lap():
    check_hour_state()
    if race["state"] != "running":
        return jsonify({"error": f"Cannot record laps. The race is {race['state']}."}), 409
    data = request.get_json(silent=True) or {}

    # Support a list of bibs for multi-lap operations via 'bibs'
    bibs = data.get('bibs')

    def _process_single(bib_val, runner_val=None):
        try:
            bib_i = int(bib_val)
        except (ValueError, TypeError):
            return {"bib": bib_val, "success": False, "error": "Invalid bib number."}

        if bib_i not in race["teams"]:
            return {"bib": bib_i, "success": False, "error": f"Bib {bib_i} does not exist."}

        team = race["teams"][bib_i]
        category = team["category"]

        # Prevent recording laps if this bib is locked as a winner
        if team.get("locked"):
            return {"bib": bib_i, "success": False, "error": "Bib is locked as category winner; modifications disabled."}

        # Prevent any laps in a category that already has a declared winner
        if category_has_winner(category):
            return {"bib": bib_i, "success": False, "error": "Category already has a declared winner; lap recording disabled."}

        runner = str(runner_val or '').upper().strip()
        if not runner:
            if category == "solo_m":
                runner = "M"
            elif category == "solo_f":
                runner = "F"
            else:
                return {"bib": bib_i, "success": False, "error": "Must specify runner M or F for Duo teams."}

        if runner not in ("M", "F"):
            return {"bib": bib_i, "success": False, "error": f"Invalid runner '{runner}'."}

        if category == "solo_m" and runner == "F":
            return {"bib": bib_i, "success": False, "error": "Solo Male bib cannot record Female laps."}
        if category == "solo_f" and runner == "M":
            return {"bib": bib_i, "success": False, "error": "Solo Female bib cannot record Male laps."}

        hour = race["current_hour"]

        if team["dnf"]:
            return {"bib": bib_i, "success": False, "error": f"Bib {bib_i} is eliminated."}

        if not team["runners"].get(runner, False):
            return {"bib": bib_i, "success": False, "error": f"Runner {runner} of Bib {bib_i} is marked DNF and cannot run."}

        now = time.time()
        clock_time = time.strftime("%H:%M:%S")

        if team["last_time_this_hour"] is None:
            lap_time = now - race["hour_start_time"]
        else:
            lap_time = now - team["last_time_this_hour"]

        max_laps = get_max_laps(category)

        if hour not in team["laps"]:
            team["laps"][hour] = []

        if len(team["laps"][hour]) >= max_laps:
            return {"bib": bib_i, "success": False, "error": f"Bib {bib_i} already completed all {max_laps} laps for Hour {hour}."}

        lap_entry = {
            "runner": runner,
            "lap_time": round(lap_time, 1),
            "ts": now,
            "lap_time_fmt": format_time(lap_time),
            "clock_time": clock_time
        }
        team["laps"][hour].append(lap_entry)
        team["last_time_this_hour"] = now

        team["runner_time"][runner] += lap_time
        team["total_time"] += lap_time

        lap_num = len(team["laps"][hour])
        add_event("lap", f"Bib {bib_i} ({runner}) · Lap {lap_num} · {format_time(lap_time)}", bib_i)

        # Check milestone: for Duo each runner must complete at least 2 laps
        try:
            runner_total_laps = sum(1 for laps in team["laps"].values() for lap in laps if lap["runner"] == runner)
        except Exception:
            runner_total_laps = 0

        if team.get("category") == "duo" and runner_total_laps >= 2 and not team.get("milestones", {}).get(runner, False):
            team.setdefault("milestones", {})[runner] = True
            add_event("milestone", f"Bib {bib_i} runner {runner} completed 2 laps", bib_i)

        return {"bib": bib_i, "success": True, "runner": runner, "hour": hour, "lap_number": lap_num, "lap_time": format_time(lap_time), "laps_remaining": max_laps - lap_num}

    # If a list of bibs provided, process each and return aggregated results
    if bibs is not None:
        if not isinstance(bibs, (list, tuple)):
            return jsonify({"error": "'bibs' must be an array of bib numbers."}), 400

        runner_global = data.get('runner')
        results = []
        for b in bibs:
            res = _process_single(b, runner_global)
            results.append(res)

        # After batch, check if hour ended
        active = [b for b, t in race["teams"].items() if not t["dnf"]]
        if len(active) <= 1:
            check_hour_state()

        return jsonify({"success": any(r.get('success') for r in results), "results": results})

    # Single bib legacy flow
    bib = data.get('bib')
    runner_in = data.get('runner')
    single_res = _process_single(bib, runner_in)

    if not single_res.get('success'):
        return jsonify({"error": single_res.get('error')}), 400

    # After single, maybe check hour
    active = [b for b, t in race["teams"].items() if not t["dnf"]]
    if len(active) <= 1:
        check_hour_state()

    return jsonify(single_res)


@app.route("/api/lap", methods=["DELETE"])
def delete_lap():
    # DELETE body may be stripped by some clients; accept JSON and delegate to core
    data = request.get_json(silent=True) or {}
    try:
        bib = int(data.get('bib', ''))
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid bib number."}), 400

    payload, status = delete_lap_core(bib, data.get('hour'), data.get('runner'))
    return (jsonify(payload), status)



def delete_lap_core(bib, hour=None, runner=None):
    """Core delete lap logic returning (payload, status_code)."""
    if bib not in race["teams"]:
        return {"error": f"Bib {bib} does not exist."}, 404

    team = race["teams"][bib]
    h = hour or race["current_hour"]
    try:
        h = int(h)
    except (ValueError, TypeError):
        return {"error": "Invalid hour."}, 400

    if h not in team["laps"] or not team["laps"][h]:
        return {"error": f"No laps recorded for Bib {bib} in Hour {h}."}, 400

    removed = None
    if runner:
        runner = str(runner).upper()
        for i in range(len(team["laps"][h]) - 1, -1, -1):
            if team["laps"][h][i]["runner"] == runner:
                removed = team["laps"][h].pop(i)
                break
        if removed is None:
            return {"error": f"No lap by runner {runner} to delete for Bib {bib} in Hour {h}."}, 400
    else:
        removed = team["laps"][h].pop()

    # adjust times
    lap_time = removed.get("lap_time", 0)
    r = removed.get("runner")
    try:
        team["runner_time"][r] = max(0, team["runner_time"].get(r, 0) - lap_time)
    except Exception:
        pass

    team["total_time"] = max(0, team.get("total_time", 0) - lap_time)

    # restore last_time_this_hour to the timestamp of the most recent remaining lap in this hour
    remaining = team["laps"].get(h, [])
    if remaining:
        last_ts = remaining[-1].get("ts")
        team["last_time_this_hour"] = last_ts
    else:
        team["last_time_this_hour"] = None

    # revoke milestone if needed
    try:
        runner_total_laps = sum(1 for laps in team["laps"].values() for lap in laps if lap["runner"] == r)
    except Exception:
        runner_total_laps = 0

    if team.get("milestones", {}).get(r, False) and runner_total_laps < 2:
        team.setdefault("milestones", {})[r] = False
        add_event("milestone_revoked", f"Bib {bib} runner {r} milestone revoked (now {runner_total_laps} laps)", bib)

    add_event("delete", f"Deleted lap for Bib {bib} ({r}) · -{format_time(lap_time)}", bib)

    return {"success": True, "bib": bib, "deleted": removed, "hour": h}, 200



@app.route("/api/lap/delete", methods=["POST"])
def delete_lap_post():
    """Fallback endpoint to delete a lap via POST (some clients do not send bodies with DELETE)."""
    data = request.get_json(silent=True) or {}
    try:
        bib = int(data.get('bib', ''))
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid bib number."}), 400

    payload, status = delete_lap_core(bib, data.get('hour'), data.get('runner'))
    return (jsonify(payload), status)


@app.route("/api/dnf", methods=["POST"])
def mark_dnf():
    check_hour_state()
    if race["state"] == "setup":
        return jsonify({"error": "Race has not started yet."}), 409

    data = request.get_json(silent=True) or {}
    try:
        bib = int(data.get("bib", ""))
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid bib number."}), 400

    if bib not in race["teams"]:
        return jsonify({"error": f"Bib {bib} does not exist."}), 404

    team = race["teams"][bib]
    category = team["category"]
    runner = str(data.get("runner", "")).upper().strip()
    
    # Infer for solo
    if not runner or runner == "BOTH":
        runner = "BOTH"
    elif runner not in ("M", "F"):
        return jsonify({"error": "Invalid runner. Must be 'M', 'F', or 'BOTH'."}), 400

    if team["dnf"]:
        return jsonify({"error": f"Bib {bib} is already eliminated."}), 409

    if runner == "BOTH":
        team["runners"]["M"] = False
        team["runners"]["F"] = False
        team["dnf"] = True
        team["dnf_info"] = {"reason": "manual", "detail": "Bib explicitly marked DNF"}
        add_event("elimination", f"Bib {bib} manually marked DNF", bib)
    else:
        if not team["runners"][runner]:
            return jsonify({"error": f"Runner {runner} of Bib {bib} is already marked DNF."}), 409
            
        team["runners"][runner] = False
        add_event("dnf", f"Runner {runner} of Bib {bib} marked DNF", bib)
        
        team["dnf"] = True
        team["dnf_info"] = {"reason": "manual", "detail": f"Runner {runner} DNF"}
        add_event("elimination", f"Bib {bib} eliminated — Runner {runner} DNF", bib)

    return jsonify({"success": True, "bib": bib, "runner": runner, "team_eliminated": True})


@app.route("/api/export", methods=["GET"])
def export_excel():
    if not race["teams"]:
        return jsonify({"error": "No data to export."}), 400

    filepath = generate_excel("race_results_Manual.xlsx")
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


@app.route("/api/race/reset", methods=["POST"])
def reset_race():
    race["state"] = "setup"
    race["current_hour"] = 1
    race["hour_start_time"] = None
    race["teams"] = {}
    race["events"] = []
    return jsonify({"success": True})


if __name__ == "__main__":
    # Show friendly startup links in the terminal so they are clickable
    host = "0.0.0.0"
    port = 5000
    try:
        import socket
        hostname = socket.gethostname()
        # Try to resolve an outward-facing IP; fall back to localhost
        try:
            local_ip = socket.gethostbyname(hostname)
        except Exception:
            local_ip = '127.0.0.1'
    except Exception:
        local_ip = '127.0.0.1'

    print(f" * Serving Flask app 'app'")
    print(f" * Debug mode: {'on' if app.debug else 'off'}")
    print(f" * Local: http://127.0.0.1:{port}/")
    if local_ip and local_ip != '127.0.0.1':
        print(f" * Network: http://{local_ip}:{port}/")
    print(f" * All interfaces: http://{host}:{port}/")

    app.run(debug=False, host=host, port=port)
