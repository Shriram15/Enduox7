import time
from openpyxl import Workbook

# -------- CONFIG --------
MAX_LAPS = 14
MAX_TIME_PER_HOUR = 3600

start_time = time.time()

# -------- TEAM SETUP --------
teams = {}
for i in range(1, 11):
    teams[i] = {
        "laps": {},
        "last_time": None,
        "total_time": 0,
        "runners": {"M": True, "F": True},
        "dnf": False,
        "current_hour": 1,
        "dnf_info": None,
        "runner_time": {"M": 0, "F": 0}
    }

def get_hour(total_time):
    return int(total_time // MAX_TIME_PER_HOUR) + 1

def format_time(sec):
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = int(sec % 60)
    return f"{h:02}:{m:02}:{s:02}"

# -------- AUTO TIME CHECK --------
def check_timeouts():
    current_time = time.time()

    for bib, team in teams.items():
        if team["dnf"]:
            continue

        total_time = current_time - start_time
        hour = get_hour(total_time)

        if hour > team["current_hour"]:
            prev_hour = team["current_hour"]
            prev_laps = team["laps"].get(prev_hour, [])

            if len(prev_laps) < MAX_LAPS:
                print(f"❌ Team {bib} AUTO ELIMINATED (did not complete {MAX_LAPS} laps in Hour {prev_hour})")
                team["dnf"] = True
                team["dnf_info"] = (prev_hour, len(prev_laps) + 1)
                continue

            team["current_hour"] = hour

# -------- WINNER CHECK --------
def check_winner():
    active = [bib for bib, t in teams.items() if not t["dnf"]]

    if len(active) == 1:
        print(f"\n🏆 WINNER: TEAM {active[0]}")
        print("🎉 RACE FINISHED 🎉")
        return True

    if len(active) == 0:
        print("\n❌ No teams left")
        return True

    return False


print("🏁 Race Started")
print("Start Time:", time.strftime("%H:%M:%S"))

# -------- INPUT LOOP --------
while True:
    check_timeouts()

    if check_winner():
        break

    choice = input("\n1: Lap | 2: DNF | 3: Quit : ")

    if choice == '3':
        break

    if choice == '1':
        choice = 'l'

    bib = int(input("Bib (1-2): "))
    if bib not in teams:
        print("Invalid bib")
        continue

    team = teams[bib]

    if team["dnf"]:
        print("Team DNF")
        continue

    if choice == '2':
        r = input("Runner (M/F): ").upper()
        team["runners"][r] = False

        if not team["runners"]["M"] and not team["runners"]["F"]:
            team["dnf"] = True
            team["dnf_info"] = ("Manual", "Both runners DNF")
            print("❌ Team eliminated")

    elif choice == 'l':
        r = input("Runner (M/F): ").upper()
        print("Press ENTER when lap completes")
        input()

        now = time.time()
        clock_time = time.strftime("%H:%M:%S")

        if team["last_time"] is None:
            lap_time = now - start_time   # first lap has no previous reference
        else:
            lap_time = now - team["last_time"]
    
        total_time = now - start_time
        team["total_time"] = total_time

        hour = get_hour(total_time)

        # Hour transition check
        if hour > team["current_hour"]:
            prev_hour = team["current_hour"]
            prev_laps = team["laps"].get(prev_hour, [])

            if len(prev_laps) < MAX_LAPS:
                print(f"❌ Team {bib} eliminated (did not complete {MAX_LAPS} laps in Hour {prev_hour})")
                team["dnf"] = True
                team["dnf_info"] = (prev_hour, len(prev_laps) + 1)
                continue

            team["current_hour"] = hour

        # Time inside hour
        hour_start = (hour - 1) * MAX_TIME_PER_HOUR
        time_in_hour = total_time - hour_start

        if time_in_hour > MAX_TIME_PER_HOUR:
            print("❌ Exceeded time limit → TEAM DNF")
            team["dnf"] = True
            team["dnf_info"] = (hour, len(team["laps"].get(hour, [])) + 1)
            continue

        if hour not in team["laps"]:
            team["laps"][hour] = []

        if len(team["laps"][hour]) >= MAX_LAPS:
            print("Max laps reached this hour")
            continue

        # Record lap
        team["laps"][hour].append({
            "runner": r,
            "lap_time": int(lap_time),
            "clock_time": clock_time
        })

        team["last_time"] = now
        team["runner_time"][r] += int(lap_time)

        print(f"✅ Team {bib} | Hour {hour} | Lap {len(team['laps'][hour])}")

    else:
        print("Invalid choice")


# -------- FINAL WINNER CHECK (if quit early) --------
active = [bib for bib, t in teams.items() if not t["dnf"]]
winner = active[0] if len(active) == 1 else None

# -------- SUMMARY --------
print("\n\n===== FINAL SUMMARY =====")

for bib, team in teams.items():
    print(f"\n--- TEAM {bib} ---")

    total_time = team["runner_time"]["M"] + team["runner_time"]["F"]
    print("Total Time:", format_time(total_time))
    print("Runner M Time:", format_time(team["runner_time"]["M"]))
    print("Runner F Time:", format_time(team["runner_time"]["F"]))

    for hour in sorted(team["laps"].keys()):
        print(f"\nHour {hour}")

        laps = team["laps"][hour]
        hour_time = sum(l["lap_time"] for l in laps)
        idle_time = max(0, MAX_TIME_PER_HOUR - hour_time)

        for i, lap in enumerate(laps, start=1):
            print(f"Lap {i} | Runner {lap['runner']} | {lap['clock_time']} | {format_time(lap['lap_time'])}")

        print(f"Idle Time: {format_time(idle_time)}")

    if team["dnf"]:
        print("Status: DNF")
        print("DNF Info:", team["dnf_info"])
    else:
        print("Status: Active")

    if winner == bib:
        print("🏆 WINNER TEAM")

print("\n\n🏆 ===== LEADERBOARD ===== 🏆")

leaderboard = []

for bib, team in teams.items():
    total_laps = sum(len(laps) for laps in team["laps"].values())
    total_time = team["runner_time"]["M"] + team["runner_time"]["F"]

    leaderboard.append({
        "bib": bib,
        "laps": total_laps,
        "time": total_time,
        "dnf": team["dnf"]
    })

# Sort:
# 1. Non-DNF first
# 2. More laps
# 3. Less time
leaderboard.sort(key=lambda x: (
    x["dnf"],          # False (0) comes before True (1)
    -x["laps"],        # more laps first
    x["time"]          # less time first
))

# Display
for rank, team in enumerate(leaderboard, start=1):
    status = "DNF" if team["dnf"] else "Active"
    rt = teams[team['bib']]["runner_time"]

    print(f"{rank}. Team {team['bib']} | Laps: {team['laps']} | "f"Time: {format_time(team['time'])} | "f"M: {format_time(rt['M'])} | F: {format_time(rt['F'])} | {status}")    

# -------- EXCEL EXPORT --------
wb = Workbook()

for bib, team in teams.items():
    ws = wb.active if bib == 1 else wb.create_sheet(f"Team_{bib}")
    ws.title = f"Team_{bib}"

    ws.cell(row=1, column=1, value="Lap")

    hours = sorted(team["laps"].keys())

    for col, hour in enumerate(hours, start=2):
        ws.cell(row=1, column=col, value=f"Hour_{hour}")

    for lap_num in range(1, MAX_LAPS + 1):
        ws.cell(row=lap_num+1, column=1, value=f"Lap {lap_num}")

        for col, hour in enumerate(hours, start=2):
            laps = team["laps"].get(hour, [])

            if team["dnf_info"] and isinstance(team["dnf_info"], tuple):
                dnf_hour, dnf_lap = team["dnf_info"]

                if hour == dnf_hour and lap_num == dnf_lap:
                    ws.cell(row=lap_num+1, column=col, value="DNF")
                    continue

            if lap_num <= len(laps):
                lap = laps[lap_num-1]
                text = f"{lap['runner']} | {lap['clock_time']} | {format_time(lap['lap_time'])}"
            else:
                text = ""

            ws.cell(row=lap_num+1, column=col, value=text)
            
            row_start = MAX_LAPS + 3

            ws.cell(row=row_start, column=1, value="Runner M Total")
            ws.cell(row=row_start, column=2, value=format_time(team["runner_time"]["M"]))

            ws.cell(row=row_start+1, column=1, value="Runner F Total")
            ws.cell(row=row_start+1, column=2, value=format_time(team["runner_time"]["F"]))

            wb.save("race_results.xlsx")

print("\n📁 Excel file saved as race_results.xlsx")